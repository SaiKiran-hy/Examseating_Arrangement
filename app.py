from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import random
import csv
import os
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///exam_seating.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
db = SQLAlchemy(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    roll_number = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    branch = db.Column(db.String(50), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    seating_arrangements = db.relationship('SeatingArrangement', backref='student', lazy=True)

class Hall(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    rows = db.Column(db.Integer, nullable=False)
    columns = db.Column(db.Integer, nullable=False)
    seating_arrangements = db.relationship('SeatingArrangement', backref='hall', lazy=True)

class Exam(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    date = db.Column(db.DateTime, nullable=False)
    exam_type = db.Column(db.String(20), nullable=False)  # 'MID' or 'SEMESTER'
    seating_arrangements = db.relationship('SeatingArrangement', backref='exam', lazy=True)

class SeatingArrangement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    exam_id = db.Column(db.Integer, db.ForeignKey('exam.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    hall_id = db.Column(db.Integer, db.ForeignKey('hall.id'), nullable=False)
    seat_row = db.Column(db.Integer, nullable=False)
    seat_column = db.Column(db.Integer, nullable=False)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xlsx', 'xls'}

def process_excel_file(file, data_type='students'):
    workbook = load_workbook(filename=file)
    worksheet = workbook.active
    headers = [str(cell.value).lower() if cell.value else '' for cell in worksheet[1]]
    data = []
    
    if data_type == 'students':
        required_headers = ['roll_number', 'name', 'branch', 'semester']
    else:  # halls
        required_headers = ['name', 'capacity', 'rows', 'columns']
    
    if not all(header in headers for header in required_headers):
        raise ValueError(f"Missing required columns. Required: {required_headers}, Found: {headers}")
    
    # Get column indices
    indices = {header: headers.index(header) for header in required_headers}
    
    for row in worksheet.iter_rows(min_row=2):
        row_data = {}
        for header, idx in indices.items():
            value = row[idx].value
            if value is not None:
                row_data[header] = str(value).strip()
        
        if all(row_data.values()):  # Check if all required fields have values
            data.append(row_data)
    
    return data

def process_csv_file(file, data_type='students'):
    content = file.read().decode('utf-8')
    file.seek(0)
    csv_reader = csv.DictReader(io.StringIO(content))
    
    if data_type == 'students':
        required_headers = ['roll_number', 'name', 'branch', 'semester']
    else:  # halls
        required_headers = ['name', 'capacity', 'rows', 'columns']
    
    if not all(header in csv_reader.fieldnames for header in required_headers):
        raise ValueError(f"Missing required columns. Required: {required_headers}, Found: {csv_reader.fieldnames}")
    
    data = []
    for row in csv_reader:
        row_data = {header: row[header].strip() for header in required_headers if row[header].strip()}
        if len(row_data) == len(required_headers):  # Check if all required fields have values
            data.append(row_data)
    
    return data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/import-data', methods=['POST'])
def import_data():
    if 'file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('manage_students'))
    
    file = request.files['file']
    data_type = request.form.get('data_type', 'students')
    
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('manage_students'))
    
    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload CSV or Excel file.', 'danger')
        return redirect(url_for('manage_students'))
    
    try:
        # Process file based on type
        if file.filename.endswith('.csv'):
            data = process_csv_file(file, data_type)
        else:
            data = process_excel_file(file, data_type)
        
        success_count = 0
        error_count = 0
        
        # Import data based on type
        if data_type == 'students':
            for item in data:
                try:
                    # Check for existing student
                    if not Student.query.filter_by(roll_number=item['roll_number']).first():
                        student = Student(
                            roll_number=item['roll_number'],
                            name=item['name'],
                            branch=item['branch'],
                            semester=int(item['semester'])
                        )
                        db.session.add(student)
                        success_count += 1
                    else:
                        error_count += 1
                except Exception as e:
                    print(f"Error adding student: {str(e)}")
                    error_count += 1
        else:  # halls
            for item in data:
                try:
                    # Check for existing hall
                    if not Hall.query.filter_by(name=item['name']).first():
                        hall = Hall(
                            name=item['name'],
                            capacity=int(item['capacity']),
                            rows=int(item['rows']),
                            columns=int(item['columns'])
                        )
                        db.session.add(hall)
                        success_count += 1
                    else:
                        error_count += 1
                except Exception as e:
                    print(f"Error adding hall: {str(e)}")
                    error_count += 1
        
        db.session.commit()
        flash(f'Successfully imported {success_count} {data_type}. {error_count} errors.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing file: {str(e)}', 'danger')
    
    return redirect(url_for('manage_students'))

@app.route('/download-template/<data_type>')
def download_template(data_type):
    workbook = Workbook()
    worksheet = workbook.active
    
    if data_type == 'students':
        headers = ['roll_number', 'name', 'branch', 'semester']
        sample_data = [
            ['CS001', 'John Doe', 'CSE', '1'],
            ['CS002', 'Jane Smith', 'CSE', '1']
        ]
    else:  # halls
        headers = ['name', 'capacity', 'rows', 'columns']
        sample_data = [
            ['Hall-A', '60', '6', '10'],
            ['Hall-B', '40', '5', '8']
        ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Add sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to memory
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{data_type}_template.xlsx'
    )

@app.route('/generate-seating/<int:exam_id>')
def generate_seating(exam_id):
    # Clear existing seating arrangements for this exam
    SeatingArrangement.query.filter_by(exam_id=exam_id).delete()
    
    exam = Exam.query.get_or_404(exam_id)
    students = Student.query.all()
    halls = Hall.query.all()
    
    if not students:
        flash('No students found. Please import student data first.', 'warning')
        return redirect(url_for('manage_exams'))
    
    if not halls:
        flash('No halls found. Please import hall data first.', 'warning')
        return redirect(url_for('manage_exams'))
    
    # Group students by branch and semester
    student_groups = {}
    for student in students:
        key = (student.branch, student.semester)
        if key not in student_groups:
            student_groups[key] = []
        student_groups[key].append(student)
    
    # Shuffle students within each group
    for group in student_groups.values():
        random.shuffle(group)
    
    # Flatten the groups while maintaining some distance between students from same branch/semester
    mixed_students = []
    max_group_size = max(len(group) for group in student_groups.values())
    
    for i in range(max_group_size):
        for group in student_groups.values():
            if i < len(group):
                mixed_students.append(group[i])
    
    # Assign seats
    current_hall_index = 0
    current_row = 0
    current_column = 0
    
    for student in mixed_students:
        if current_hall_index >= len(halls):
            flash('Not enough space in halls for all students', 'warning')
            break
            
        hall = halls[current_hall_index]
        
        # Create seating arrangement
        arrangement = SeatingArrangement(
            exam_id=exam_id,
            student_id=student.id,
            hall_id=hall.id,
            seat_row=current_row,
            seat_column=current_column
        )
        db.session.add(arrangement)
        
        # Move to next position (skip one seat for spacing)
        current_column += 2
        if current_column >= hall.columns:
            current_column = 0
            current_row += 1
            
        if current_row >= hall.rows:
            current_row = 0
            current_hall_index += 1
    
    db.session.commit()
    flash('Seating arrangement generated successfully!', 'success')
    return redirect(url_for('view_seating', exam_id=exam_id))

@app.route('/view_seating/<int:exam_id>')
def view_seating(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    arrangements = SeatingArrangement.query.filter_by(exam_id=exam_id).all()
    halls = Hall.query.all()
    
    # Organize seating data by hall
    hall_seating = {}
    for hall in halls:
        hall_seating[hall.id] = {
            'name': hall.name,
            'rows': hall.rows,
            'columns': hall.columns,
            'seats': [[None for _ in range(hall.columns)] for _ in range(hall.rows)]
        }
    
    # Fill in student information
    for arrangement in arrangements:
        hall_id = arrangement.hall_id
        row = arrangement.seat_row
        col = arrangement.seat_column
        student = arrangement.student
        
        hall_seating[hall_id]['seats'][row][col] = {
            'roll_number': student.roll_number,
            'name': student.name,
            'branch': student.branch,
            'semester': student.semester
        }
    
    return render_template('seating.html', 
                         exam=exam, 
                         hall_seating=hall_seating,
                         halls=halls)

@app.route('/students', methods=['GET', 'POST'])
def manage_students():
    if request.method == 'POST':
        try:
            student = Student(
                roll_number=request.form['roll_number'],
                name=request.form['name'],
                branch=request.form['branch'],
                semester=int(request.form['semester'])
            )
            db.session.add(student)
            db.session.commit()
            flash('Student added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error adding student. Please check if the roll number is unique.', 'danger')
        return redirect(url_for('manage_students'))
    
    students = Student.query.all()
    return render_template('students.html', students=students)

@app.route('/halls', methods=['GET', 'POST'])
def manage_halls():
    if request.method == 'POST':
        hall = Hall(
            name=request.form['name'],
            capacity=int(request.form['capacity']),
            rows=int(request.form['rows']),
            columns=int(request.form['columns'])
        )
        db.session.add(hall)
        db.session.commit()
        return redirect(url_for('manage_halls'))
    
    halls = Hall.query.all()
    return render_template('halls.html', halls=halls)

@app.route('/exams', methods=['GET', 'POST'])
def manage_exams():
    if request.method == 'POST':
        exam = Exam(
            name=request.form['name'],
            date=datetime.strptime(request.form['date'], '%Y-%m-%d'),
            exam_type=request.form['exam_type']
        )
        db.session.add(exam)
        db.session.commit()
        return redirect(url_for('manage_exams'))
    
    exams = Exam.query.all()
    return render_template('exams.html', exams=exams)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)